#!/usr/bin/env python3
"""
Reads "SleepOver - Email Signatures.xlsx" and writes employees.json, the single
data source signature-generator.js consumes.

Why two steps: the workbook is the thing the client edits and the thing that
mirrors HubSpot, but Node cannot read .xlsx without a dependency. So Python
flattens it to JSON and the Node generator stays dependency-free.

Run:  python3 build-data.py  &&  node signature-generator.js
"""

import json
import os
import re
import struct
from urllib.parse import quote

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
XLSX = os.path.join(ROOT, "SleepOver - Email Signatures.xlsx")
ASSETS = os.path.join(ROOT, "assets")
OUT = os.path.join(HERE, "employees.json")

# Sheet -> (asset sub-folder, label used in the logo alt text)
SHEETS = {
    "SLEEPOVER INTERNATIONAL": ("International", "SleepOver International"),
    "SLEEPOVER NIGERIA":       ("Nigeria",       "SleepOver Nigeria"),
    "SLEEPOVER TANZANIA":      ("Tanzania",      "SleepOver Tanzania"),
    "SLEEPOVER BOTSWANA":      ("Botswana",      "SleepOver Botswana"),
    "SLEEPOVER NAMIBIA":       ("Namibia",       "SleepOver Namibia"),
}

# Columns, 1-indexed, as laid out in the workbook.
C_NAME, C_NAME_URL, C_POSITION, C_PHONE, C_EMAIL = 1, 2, 3, 4, 5
C_ADDRESS, C_OFFICE_PHONE, C_WEBSITE, C_REGION = 6, 7, 8, 9

# Layout limits, in CSS px at 1x. The name artwork is placed at half its export
# size so it stays sharp on retina; NAME_MAX_W is the usable width of the left
# column (380 - 15 left padding - 22 gutter), so a very long name scales down
# instead of pushing the divider out of the card.
NAME_SCALE = 0.5
NAME_MAX_W = 343
LOGO_W = 150

# Where the generated pages are published. Each person's page is
# PAGES_BASE + "<Their-Name>.html"; the repo layout has not changed, so this is
# the same shape as the old links.
PAGES_BASE = "https://gabrielmagcr.github.io/SleepOverSignatures/signature-generator/"

# The one office with a verified Google Maps pin. Everything else gets a maps
# search on the address text, which resolves correctly but is not a saved pin.
FOURWAYS_PIN = "https://maps.app.goo.gl/4YZ1YCxyxTXiUDC57"


def png_size(path):
    """Width/height straight out of the PNG IHDR. Avoids a Pillow dependency."""
    with open(path, "rb") as fh:
        head = fh.read(24)
    if head[:8] != b"\x89PNG\r\n\x1a\n":
        raise ValueError("not a PNG: " + path)
    return struct.unpack(">II", head[16:24])


def wrap_address(raw):
    """
    Address lines for the signature.

    Newlines in the cell are the author's own line breaks and are always
    honoured. A single-line cell (Nigeria) is wrapped on its commas into
    ~30-character lines so it matches the shape of the hand-broken ones instead
    of reflowing inside the 188px column.
    """
    raw = (raw or "").strip()
    if not raw:
        return [], False

    if "\n" in raw:
        return [l.strip() for l in raw.split("\n") if l.strip()], False

    parts = [p.strip() for p in raw.split(",") if p.strip()]
    lines, cur = [], ""
    for i, part in enumerate(parts):
        piece = part + ("," if i < len(parts) - 1 else "")
        if cur and len(cur) + 1 + len(piece) > 30:
            lines.append(cur)
            cur = piece
        else:
            cur = (cur + " " + piece).strip()
    if cur:
        lines.append(cur)
    return lines, True


def norm_phone(raw):
    """
    Office phone as it should read. Piers' row is stored as '27(0)10 110 9910'
    with a trailing newline while every other International row reads
    '+27 (0)10 110 9910'; this makes the 18 of them identical.
    """
    v = re.sub(r"\s+", " ", (raw or "").strip())
    if not v:
        return ""
    if not v.startswith("+"):
        v = "+" + v
    return re.sub(r"^\+(\d+)\(", r"+\1 (", v)


def maps_url(lines):
    joined = ", ".join(lines)
    if "Little Fourways" in joined:
        return FOURWAYS_PIN
    return "https://www.google.com/maps/search/?api=1&query=" + quote(joined)


def main():
    import openpyxl

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    people, flags = [], []

    def flag(kind, name, region, detail=""):
        """One structured note. The generator groups these by kind so the
        master page shows three lines instead of one per person."""
        flags.append({"kind": kind, "name": name, "region": region, "detail": detail})

    for sheet, (folder, region_label) in SHEETS.items():
        ws = wb[sheet]

        for row in range(3, ws.max_row + 1):
            name = ws.cell(row, C_NAME).value
            if not name or not str(name).strip():
                continue
            name = str(name).strip()

            email = str(ws.cell(row, C_EMAIL).value or "").strip()
            name_url = str(ws.cell(row, C_NAME_URL).value or "").strip()
            if not (email and name_url):
                flag("missing-data", name, folder, "no email or name URL in the workbook")
                continue

            # Name artwork: measured locally, placed at half size, then capped
            # to the column width so long names cannot break the grid.
            art = os.path.join(ASSETS, "SO Signatures Names", folder, name + ".png")
            aw, ah = png_size(art)
            scale = min(NAME_SCALE, NAME_MAX_W / aw)
            name_w, name_h = round(aw * scale), round(ah * scale)
            if scale < NAME_SCALE:
                flag("artwork-scaled", name, folder,
                     f"{aw}px wide, placed at {name_w}px, "
                     f"{round((1 - scale / NAME_SCALE) * 100)}% smaller than the rest")

            # Region lockup, scaled to a fixed width off its real aspect ratio.
            logo_file = f"SO_{folder}.png"
            lw, lh = png_size(os.path.join(ASSETS, "Regions", logo_file))
            logo_h = round(LOGO_W * lh / lw)

            address, wrapped = wrap_address(ws.cell(row, C_ADDRESS).value)
            if wrapped:
                flag("address-wrapped", name, folder)

            office_phone = norm_phone(ws.cell(row, C_OFFICE_PHONE).value)
            if not office_phone:
                flag("no-office-phone", name, folder)

            website = str(ws.cell(row, C_WEBSITE).value or "sleepover.travel").strip()

            slug = re.sub(r"\s+", "-", name)

            people.append({
                "name": name,
                "slug": slug,
                "pageUrl": PAGES_BASE + quote(slug) + ".html",
                "position": str(ws.cell(row, C_POSITION).value or "").strip().upper(),
                "phone": str(ws.cell(row, C_PHONE).value or "").strip(),
                "email": email,
                "region": str(ws.cell(row, C_REGION).value or folder).strip().upper(),
                "regionLabel": region_label,
                "regionFolder": folder,
                "sheet": sheet,
                "nameImage": {"url": name_url, "w": name_w, "h": name_h},
                "logo": {
                    "url": f"https://magneticcreative.com/hubfs/SleepOver%20Signature/regions/{quote(logo_file)}",
                    "w": LOGO_W,
                    "h": logo_h,
                    "alt": region_label,
                },
                "officePhone": office_phone,
                "website": website,
                "websiteUrl": "https://" + website.replace("https://", "").replace("http://", ""),
                "address": address,
                "mapsUrl": maps_url(address),
            })

    with open(OUT, "w", encoding="utf8") as fh:
        json.dump({"people": people, "flags": flags}, fh, indent=2, ensure_ascii=False)

    by_kind = {}
    for f in flags:
        by_kind.setdefault(f["kind"], []).append(f["name"])
    print(f"employees.json: {len(people)} people, {len(flags)} notes")
    for kind, names in by_kind.items():
        print(f"  - {kind}: {len(names)} ({', '.join(names[:3])}{'...' if len(names) > 3 else ''})")


if __name__ == "__main__":
    main()
