#!/usr/bin/env python3
"""
Writes each person's published signature-page URL into column K of
"../SleepOver - Email Signatures.xlsx", so the workbook is the one place that
holds both the HubSpot image URLs and the link you send the person.

Run after build-data.py, which is where PAGES_BASE lives:

    python3 build-data.py && python3 write-page-urls.py

Safe to re-run: it matches on the name in column A and overwrites column K.
"""

import json
import os
import sys
from copy import copy

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(os.path.dirname(HERE), "SleepOver - Email Signatures.xlsx")
JSON = os.path.join(HERE, "employees.json")

HEADER = "Signature Page URL"
COL = 11  # K


def main():
    with open(JSON, encoding="utf8") as fh:
        people = json.load(fh)["people"]
    by_sheet = {}
    for p in people:
        by_sheet.setdefault(p["sheet"], {})[p["name"]] = p["pageUrl"]

    wb = openpyxl.load_workbook(XLSX)
    written, missing = 0, []

    for sheet, urls in by_sheet.items():
        ws = wb[sheet]

        # Header, styled like the columns beside it.
        head = ws.cell(1, COL)
        head.value = HEADER
        head._style = copy(ws.cell(1, COL - 1)._style)

        seen = set()
        for row in range(3, ws.max_row + 1):
            name = ws.cell(row, COL - 10).value  # column A
            if not name:
                continue
            name = str(name).strip()
            if name not in urls:
                continue
            cell = ws.cell(row, COL)
            cell.value = urls[name]
            cell.hyperlink = urls[name]
            cell._style = copy(ws.cell(row, 2)._style)  # match the Name URL column
            seen.add(name)
            written += 1
        missing += [n for n in urls if n not in seen]

    if missing:
        print("No matching row for:", ", ".join(missing), file=sys.stderr)
        return 1

    ws_any = wb[next(iter(by_sheet))]
    ws_any.column_dimensions[ws_any.cell(1, COL).column_letter].width = 42

    wb.save(XLSX)
    print(f"Wrote column K ({HEADER}) for {written} people across {len(by_sheet)} sheets.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
